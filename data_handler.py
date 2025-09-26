# Data handling and export for LeadSprinter
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

class DataHandler:
    def __init__(self):
        self.results_df = pd.DataFrame()
        self.search_metadata = {}
        
    def store_results(self, data):
        """Store results in DataFrame"""
        if isinstance(data, pd.DataFrame):
            self.results_df = data.copy()
        elif isinstance(data, list):
            self.results_df = pd.DataFrame(data)
        else:
            raise ValueError("Data must be a pandas DataFrame or list of dictionaries")
        
        # Clean and validate data
        self.clean_data()
    
    def clean_data(self):
        """Clean and standardize the data"""
        if self.results_df.empty:
            return
        
        # Remove duplicates
        self.results_df = self.results_df.drop_duplicates(subset=['linkedin_url'], keep='first')
        
        # Clean text fields
        text_columns = ['name', 'title', 'company', 'location']
        for col in text_columns:
            if col in self.results_df.columns:
                self.results_df[col] = self.results_df[col].astype(str).str.strip()
                self.results_df[col] = self.results_df[col].replace('nan', 'N/A')
                self.results_df[col] = self.results_df[col].replace('', 'N/A')
        
        # Validate LinkedIn URLs
        if 'linkedin_url' in self.results_df.columns:
            self.results_df = self.results_df[
                self.results_df['linkedin_url'].str.contains('linkedin.com', na=False)
            ]
        
        # Add timestamp
        self.results_df['scraped_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Reset index
        self.results_df.reset_index(drop=True, inplace=True)
    
    def add_search_metadata(self, search_params):
        """Add search parameters as metadata"""
        self.search_metadata = {
            'search_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'job_titles': search_params.get('-JOB_TITLES-', ''),
            'locations': search_params.get('-LOCATIONS-', ''),
            'num_results_requested': search_params.get('-NUM_RESULTS-', 0),
            'industry_filter': search_params.get('-INDUSTRY-', 'All Industries'),
            'company_size_filter': search_params.get('-COMPANY_SIZE-', 'All Sizes'),
            'results_found': len(self.results_df)
        }
    
    def get_summary_stats(self):
        """Get summary statistics of the data"""
        if self.results_df.empty:
            return {}
        
        stats = {
            'total_profiles': len(self.results_df),
            'profiles_with_email': len(self.results_df[self.results_df['email'].notna()]),
            'unique_companies': self.results_df['company'].nunique(),
            'unique_locations': self.results_df['location'].nunique(),
            'top_companies': self.results_df['company'].value_counts().head(5).to_dict(),
            'top_locations': self.results_df['location'].value_counts().head(5).to_dict(),
            'email_rate': f"{(len(self.results_df[self.results_df['email'].notna()]) / len(self.results_df) * 100):.1f}%"
        }
        
        return stats
    
    def export_to_excel(self, filename=None, search_params=None):
        """Export DataFrame to Excel with professional formatting"""
        if self.results_df.empty:
            raise ValueError("No data to export")
        
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"LeadSprinter_Results_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Create full file path
        filepath = os.path.abspath(filename)
        
        # Add search metadata if provided
        if search_params:
            self.add_search_metadata(search_params)
        
        # Create workbook and worksheet
        wb = Workbook()
        
        # Main results sheet
        ws_results = wb.active
        ws_results.title = "Lead Results"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Prepare data for export
        export_columns = ['name', 'title', 'company', 'location', 'linkedin_url', 'email', 'scraped_date']
        export_df = self.results_df[export_columns].copy()
        
        # Rename columns for better presentation
        column_names = {
            'name': 'Full Name',
            'title': 'Job Title',
            'company': 'Company',
            'location': 'Location',
            'linkedin_url': 'LinkedIn Profile',
            'email': 'Email Address',
            'scraped_date': 'Date Scraped'
        }
        export_df.rename(columns=column_names, inplace=True)
        
        # Write headers
        for col_idx, column_name in enumerate(export_df.columns, 1):
            cell = ws_results.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Write data
        for row_idx, row_data in enumerate(dataframe_to_rows(export_df, index=False, header=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_results.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border_style
                
                # Special formatting for LinkedIn URLs
                if col_idx == 5 and value:  # LinkedIn URL column
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")
        
        # Auto-adjust column widths
        for column in ws_results.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws_results.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        if self.search_metadata:
            ws_summary = wb.create_sheet("Search Summary")
            
            # Summary data
            summary_data = [
                ["Search Summary", ""],
                ["Search Date", self.search_metadata.get('search_date', 'N/A')],
                ["Job Titles Searched", self.search_metadata.get('job_titles', 'N/A')],
                ["Locations Searched", self.search_metadata.get('locations', 'N/A')],
                ["Results Requested", self.search_metadata.get('num_results_requested', 'N/A')],
                ["Results Found", self.search_metadata.get('results_found', 'N/A')],
                ["Industry Filter", self.search_metadata.get('industry_filter', 'N/A')],
                ["Company Size Filter", self.search_metadata.get('company_size_filter', 'N/A')],
                ["", ""],
                ["Statistics", ""],
            ]
            
            # Add statistics
            stats = self.get_summary_stats()
            for key, value in stats.items():
                if key not in ['top_companies', 'top_locations']:
                    formatted_key = key.replace('_', ' ').title()
                    summary_data.append([formatted_key, value])
            
            # Write summary data
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                ws_summary.cell(row=row_idx, column=2, value=value)
            
            # Auto-adjust summary sheet columns
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 40
        
        # Save workbook
        try:
            wb.save(filepath)
            return filepath
        except Exception as e:
            raise Exception(f"Failed to save Excel file: {str(e)}")
    
    def export_to_csv(self, filename=None):
        """Export DataFrame to CSV"""
        if self.results_df.empty:
            raise ValueError("No data to export")
        
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"LeadSprinter_Results_{timestamp}.csv"
        
        # Ensure .csv extension
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        # Create full file path
        filepath = os.path.abspath(filename)
        
        # Export to CSV
        try:
            self.results_df.to_csv(filepath, index=False)
            return filepath
        except Exception as e:
            raise Exception(f"Failed to save CSV file: {str(e)}")
    
    def has_data(self):
        """Check if handler has data"""
        return not self.results_df.empty
    
    def get_data(self):
        """Get the stored DataFrame"""
        return self.results_df.copy()
    
    def clear_data(self):
        """Clear stored data"""
        self.results_df = pd.DataFrame()
        self.search_metadata = {}

# Legacy functions for backward compatibility
def store_results(data):
    handler = DataHandler()
    handler.store_results(data)
    return handler

def export_to_excel(df, filename):
    handler = DataHandler()
    handler.store_results(df)
    return handler.export_to_excel(filename)
